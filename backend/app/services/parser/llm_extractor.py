"""
LLM extraction using Google Gemini.
Provides three modes:
  - extract_all(markdown): text-based extraction from pre-converted markdown
  - extract_all_from_file(gemini_file_uri, selected_sheets): native Gemini File API extraction
    that preserves images, charts, and visual content from office files
  - extract_document_metadata(markdown, filename): lightweight metadata-only pass
"""
import ast
import json
import re
import logging
import traceback
import google.generativeai as genai
from app.core.config import get_settings
from app.services.parser.prompts_store import get_prompt

logger = logging.getLogger(__name__)

_client_initialized = False


def _parse_llm_json(raw: str) -> dict:
    """
    Parse LLM JSON output with fallbacks for common Gemini quirks:
      - Strips markdown code fences
      - Fixes Python literals (None/True/False → null/true/false)
      - Falls back to ast.literal_eval for single-quoted keys/values
    Raises json.JSONDecodeError if all attempts fail.
    """
    raw = raw.strip()

    # Strip markdown code fences (```json ... ``` or ``` ... ```)
    if raw.startswith("```"):
        lines = raw.splitlines()
        end = -1 if lines[-1].strip() == "```" else len(lines)
        raw = "\n".join(lines[1:end]).strip()

    # Attempt 1: strict JSON
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass

    # Attempt 2: fix Python literals then retry
    fixed = re.sub(r'\bNone\b', 'null', raw)
    fixed = re.sub(r'\bTrue\b', 'true', fixed)
    fixed = re.sub(r'\bFalse\b', 'false', fixed)
    try:
        return json.loads(fixed)
    except json.JSONDecodeError:
        pass

    # Attempt 3: ast.literal_eval handles single-quoted keys/values
    try:
        result = ast.literal_eval(raw)
        if isinstance(result, (dict, list)):
            # Round-trip through json to normalise types
            return json.loads(json.dumps(result))
    except Exception:
        pass

    # Re-raise with original raw for diagnostics
    return json.loads(raw)


def _init_client():
    global _client_initialized
    if not _client_initialized:
        s = get_settings()
        genai.configure(api_key=s.gemini_api_key)
        _client_initialized = True


# Prompts are now managed via prompts_store.py (editable at runtime via /admin)


# ─── Core LLM call ────────────────────────────────────────────

def _call_llm(prompt: str) -> dict:
    _init_client()
    s = get_settings()
    model = genai.GenerativeModel(
        model_name=s.gemini_model,
        system_instruction=get_prompt("system"),
        generation_config=genai.GenerationConfig(
            response_mime_type="application/json",
            temperature=0.1,
        ),
    )
    response = model.generate_content(prompt)
    return _parse_llm_json(response.text)


# ─── Public extraction functions ──────────────────────────────

def extract_all(markdown: str) -> dict:
    """
    Combined single-pass extraction: returns flows with APIs nested inside.
    Output shape: { "flows": [ { name, description, steps, apis } ] }
    On failure returns { "flows": [], "_error": "..." }
    """
    prompt = get_prompt("extract_all").replace("{content}", markdown)
    try:
        result = _call_llm(prompt)
        if "flows" not in result:
            result = {"flows": [], "_error": "LLM did not return a flows key"}
        return result
    except Exception as e:
        return {"flows": [], "_error": str(e)}


def extract_all_from_file(
    gemini_file_uri: str,
    selected_sheets: list[str] | None = None,
    sheet_kinds: dict[str, str] | None = None,
    flow_sequence: dict[str, list[dict]] | None = None,
) -> dict:
    """
    Extract using Gemini File API — the raw file is passed directly so Gemini
    can read embedded images, charts, and diagrams that markitdown would drop.

    selected_sheets: if provided, Gemini is instructed to focus only on those sheets.
    Output shape: { "flows": [ { name, description, steps, apis } ] }
    On failure returns { "flows": [], "_error": "..." }
    """
    _init_client()
    s = get_settings()

    sheet_hint = ""
    if selected_sheets:
        kind_descriptions = {
            "api_spec":   "contains API endpoint definitions (URL, request/response fields)",
            "error_code": "reference table of result/error codes",
            "edge_case":  "runtime handling logic (retry, inquiry, timeouts)",
            "mapping":    "code mapping or lookup table — extract only if referenced by an API",
            "flow":       "flow diagram or sequence overview — use to understand the overall process",
            "metadata":   "changelog, environment info, overview — informational only",
        }
        lines = ["Sheet context for this document:"]
        for name in selected_sheets:
            kind = (sheet_kinds or {}).get(name, "api_spec")
            desc = kind_descriptions.get(kind, "")
            lines.append(f'  - "{name}": {kind.replace("_", " ").title()} — {desc}')
        lines.append("\nFocus ONLY on the sheets listed above. Ignore all other sheets.")
        sheet_hint = "\n" + "\n".join(lines) + "\n"

    # Append user-defined flow step sequences as structured hints
    if flow_sequence:
        seq_lines = [
            "\nFlow sequence hints (user-defined — follow these step orders when building flow_step arrays):"
        ]
        for flow_name, steps in flow_sequence.items():
            seq_lines.append(f'  Flow "{flow_name}":')
            for i, step in enumerate(steps, 1):
                seq_lines.append(f'    Step {i}: "{step["sheet_name"]}" sheet — {step.get("label", step["sheet_name"])}')
        sheet_hint += "\n".join(seq_lines) + "\n"

    prompt = get_prompt("extract_all_file").replace("{sheet_hint}", sheet_hint)

    try:
        logger.info("extract_all_from_file: fetching Gemini file uri=%s", gemini_file_uri)
        # genai.get_file expects just the file name (e.g. "files/xxx" or "xxx"),
        # not the full URI. Extract the name portion if a full URL was stored.
        file_name = gemini_file_uri
        if "generativelanguage.googleapis.com" in gemini_file_uri:
            # Extract path after the domain, e.g. "/v1beta/files/5gnr3pdfkz0a" -> "files/5gnr3pdfkz0a"
            from urllib.parse import urlparse
            parsed = urlparse(gemini_file_uri)
            # path is like /v1beta/files/<id>
            parts = parsed.path.lstrip("/").split("/")
            # find "files" segment
            try:
                idx = parts.index("files")
                file_name = "/".join(parts[idx:])
            except ValueError:
                file_name = parts[-1]
        uploaded_file = genai.get_file(file_name)
        logger.info("Gemini file state=%s size=%s", getattr(uploaded_file, "state", "?"), getattr(uploaded_file, "size_bytes", "?"))
        model = genai.GenerativeModel(
            model_name=s.gemini_model,
            system_instruction=get_prompt("system"),
            generation_config=genai.GenerationConfig(
                response_mime_type="application/json",
                temperature=0.1,
            ),
        )
        logger.info("Sending generate_content to Gemini model=%s", s.gemini_model)
        response = model.generate_content([uploaded_file, prompt])
        raw = response.text
        logger.info("Gemini response received, raw length=%d chars", len(raw))
        result = _parse_llm_json(raw)
        if isinstance(result, list):
            result = {"flows": result}
        if "flows" not in result:
            logger.warning("Gemini returned JSON without 'flows' key: %s", list(result.keys()))
            result = {"flows": [], "_error": "LLM did not return a flows key"}
        else:
            logger.info("Extraction complete: %d flow(s) returned", len(result["flows"]))
        return result
    except Exception as e:
        logger.error("extract_all_from_file failed uri=%s\n%s", gemini_file_uri, traceback.format_exc())
        return {"flows": [], "_error": str(e)}


def extract_all_from_xlsx(
    file_path: str,
    selected_sheets: list[str] | None = None,
    sheet_kinds: dict[str, str] | None = None,
    flow_sequence: dict[str, list[dict]] | None = None,
    parser: str = "openpyxl",
) -> dict:
    """
    Extract from an XLSX file by:
      1. Converting selected sheets to markdown tables (text)
      2. Extracting embedded images from those sheets and uploading to Gemini
      3. Calling generate_content with [*images, text, prompt+hints]

    This preserves embedded diagrams/charts while Gemini File API's XLSX MIME
    type is unsupported.
    """
    import openpyxl
    import io

    _init_client()
    s = get_settings()

    # ── Build sheet/flow hints (same logic as extract_all_from_file) ──
    sheet_hint = ""
    if selected_sheets:
        kind_descriptions = {
            "api_spec":   "contains API endpoint definitions (URL, request/response fields)",
            "error_code": "reference table of result/error codes",
            "edge_case":  "runtime handling logic (retry, inquiry, timeouts)",
            "mapping":    "code mapping or lookup table — extract only if referenced by an API",
            "flow":       "flow diagram or sequence overview — use to understand the overall process",
            "metadata":   "changelog, environment info, overview — informational only",
        }
        lines = ["Sheet context for this document:"]
        for name in selected_sheets:
            kind = (sheet_kinds or {}).get(name, "api_spec")
            desc = kind_descriptions.get(kind, "")
            lines.append(f'  - "{name}": {kind.replace("_", " ").title()} — {desc}')
        lines.append("\nFocus ONLY on the sheets listed above. Ignore all other sheets.")
        sheet_hint = "\n" + "\n".join(lines) + "\n"

    if flow_sequence:
        seq_lines = [
            "\nFlow sequence hints (user-defined — follow these step orders when building flow_step arrays):"
        ]
        for flow_name, steps in flow_sequence.items():
            seq_lines.append(f'  Flow "{flow_name}":')
            for i, step in enumerate(steps, 1):
                seq_lines.append(f'    Step {i}: "{step["sheet_name"]}" sheet — {step.get("label", step["sheet_name"])}')
        sheet_hint += "\n".join(seq_lines) + "\n"

    prompt = get_prompt("extract_all_file").replace("{sheet_hint}", sheet_hint)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = selected_sheets if selected_sheets else wb.sheetnames

        # ── Convert selected sheets to markdown ──
        if parser == "markitdown":
            from markitdown import MarkItDown
            sheet_text = MarkItDown().convert(file_path).text_content
        else:
            # openpyxl (default): handles merged cells, multiple tables per sheet,
            # sparse rows as prose.
            from app.services.parser.ingestion import xlsx_to_markdown
            sheet_text = xlsx_to_markdown(file_path, sheet_names)

        # ── Extract embedded images from selected sheets ──
        uploaded_images = []
        for name in sheet_names:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for img in getattr(ws, "_images", []):
                try:
                    img_bytes = img._data()
                    # Detect mime type from magic bytes
                    if img_bytes[:4] == b'\x89PNG':
                        mime = "image/png"
                        ext = "png"
                    elif img_bytes[:2] == b'\xff\xd8':
                        mime = "image/jpeg"
                        ext = "jpg"
                    else:
                        mime = "image/png"
                        ext = "png"
                    gfile = genai.upload_file(
                        path=io.BytesIO(img_bytes),
                        mime_type=mime,
                        display_name=f"{name}_image.{ext}",
                    )
                    uploaded_images.append(gfile)
                    logger.info("Uploaded embedded image from sheet=%s mime=%s", name, mime)
                except Exception:
                    logger.warning("Failed to upload image from sheet=%s", name, exc_info=True)

        wb.close()

        logger.info(
            "extract_all_from_xlsx: %d sheet(s), %d embedded image(s)",
            len(sheet_names), len(uploaded_images),
        )

        model = genai.GenerativeModel(
            model_name=s.gemini_model,
            system_instruction=get_prompt("system"),
            generation_config=genai.GenerationConfig(
                response_mime_type="application/json",
                temperature=0.1,
            ),
        )

        # Build content: images first, then sheet text, then prompt
        content = [*uploaded_images, sheet_text, prompt]
        response = model.generate_content(content)
        raw = response.text
        logger.info("Gemini response received, raw length=%d chars", len(raw))
        logger.debug("sheet_text sent to Gemini (first 2000 chars):\n%s", sheet_text[:2000])
        logger.debug("Gemini raw response (first 2000 chars):\n%s", raw[:2000])
        result = _parse_llm_json(raw)
        if isinstance(result, list):
            result = {"flows": result}
        if "flows" not in result:
            logger.warning("Gemini returned JSON without 'flows' key: %s", list(result.keys()))
            result = {"flows": [], "_error": "LLM did not return a flows key"}
        else:
            logger.info("Extraction complete: %d flow(s) returned", len(result["flows"]))
        # Attach generated sheet markdown so callers can store it for review/editing
        result["_sheet_markdown"] = sheet_text
        return result
    except Exception as e:
        logger.error("extract_all_from_xlsx failed path=%s\n%s", file_path, traceback.format_exc())
        return {"flows": [], "_error": str(e)}


def run_playground(
    file_bytes: bytes,
    filename: str,
    sheet_selection: dict | None = None,
    flow_sequence: dict | None = None,
) -> dict:
    """
    Run a full extraction and return every intermediate step for debugging.
    sheet_selection: { "selected_sheets": [...], "sheet_kinds": {...} }
    flow_sequence:   { flow_name: [{"sheet_name": ..., "label": ...}] }
    Returns: { "steps": [{"label", "type", "content"}], "error": str|None }
    """
    import io as _io
    import os
    import tempfile

    steps: list[dict] = []

    def _add(label: str, type_: str, content):
        steps.append({"label": label, "type": type_, "content": content})

    try:
        _init_client()
        s = get_settings()

        # ── Normalise sheet_selection ──
        # Supports both the UI-saved format {"selected_sheets", "sheet_kinds"}
        # and the exported file format {"selected", "kinds"}
        sel = sheet_selection or {}
        selected_sheets = sel.get("selected_sheets") or sel.get("selected") or None
        sheet_kinds = sel.get("sheet_kinds") or sel.get("kinds") or None

        # ── Normalise flow_sequence ──
        # Supports both the pipeline format {flow_name: [{"sheet_name", "label"}]}
        # and the exported file format {"flows": [...], "sequences": {flow_name: [{"sheetName", "label"}]}}
        norm_flow_sequence: dict | None = None
        if flow_sequence:
            raw_seq = flow_sequence.get("sequences") if "sequences" in flow_sequence else flow_sequence
            norm_flow_sequence = {}
            for flow_name, fsteps in raw_seq.items():
                norm_steps = []
                for step in fsteps:
                    if isinstance(step, str):
                        norm_steps.append({"sheet_name": step, "label": step})
                    else:
                        # camelCase sheetName → snake_case sheet_name
                        sheet = step.get("sheet_name") or step.get("sheetName", "")
                        norm_steps.append({"sheet_name": sheet, "label": step.get("label", sheet)})
                norm_flow_sequence[flow_name] = norm_steps

        # ── Build hints string ──
        kind_descriptions = {
            "api_spec":   "contains API endpoint definitions (URL, request/response fields)",
            "error_code": "reference table of result/error codes",
            "edge_case":  "runtime handling logic (retry, inquiry, timeouts)",
            "mapping":    "code mapping or lookup table",
            "flow":       "flow diagram or sequence overview",
            "metadata":   "changelog, environment info, overview",
        }
        sheet_hint = ""
        if selected_sheets:
            lines = ["Sheet context for this document:"]
            for name in selected_sheets:
                kind = (sheet_kinds or {}).get(name, "api_spec")
                desc = kind_descriptions.get(kind, "")
                lines.append(f'  - "{name}": {kind.replace("_", " ").title()} — {desc}')
            lines.append("\nFocus ONLY on the sheets listed above. Ignore all other sheets.")
            sheet_hint = "\n" + "\n".join(lines) + "\n"

        if norm_flow_sequence:
            seq_lines = ["\nFlow sequence hints (user-defined):"]
            for flow_name, fsteps in norm_flow_sequence.items():
                seq_lines.append(f'  Flow "{flow_name}":')
                for i, step in enumerate(fsteps, 1):
                    seq_lines.append(
                        f'    Step {i}: "{step["sheet_name"]}" — {step["label"]}'
                    )
            sheet_hint += "\n".join(seq_lines) + "\n"

        _add(
            "Sheet / Flow Hints",
            "text",
            sheet_hint.strip() if sheet_hint.strip() else "(none — all sheets, no flow sequence)",
        )

        is_xlsx = filename.lower().endswith(".xlsx")

        if is_xlsx:
            import openpyxl
            from app.services.parser.ingestion import xlsx_to_markdown

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = tmp.name

            try:
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                sheet_names = selected_sheets if selected_sheets else list(wb.sheetnames)

                sheet_text = xlsx_to_markdown(tmp_path, sheet_names)
                _add("Sheet Markdown (XLSX → text)", "markdown", sheet_text)

                uploaded_images = []
                image_log = []
                for sname in sheet_names:
                    if sname not in wb.sheetnames:
                        continue
                    ws = wb[sname]
                    for img in getattr(ws, "_images", []):
                        try:
                            img_bytes = img._data()
                            if img_bytes[:4] == b'\x89PNG':
                                mime, ext = "image/png", "png"
                            elif img_bytes[:2] == b'\xff\xd8':
                                mime, ext = "image/jpeg", "jpg"
                            else:
                                mime, ext = "image/png", "png"
                            gfile = genai.upload_file(
                                path=_io.BytesIO(img_bytes),
                                mime_type=mime,
                                display_name=f"{sname}_image.{ext}",
                            )
                            uploaded_images.append(gfile)
                            image_log.append(
                                f"Sheet '{sname}': {mime} ({len(img_bytes):,} bytes) → {gfile.name}"
                            )
                        except Exception as ex:
                            image_log.append(f"Sheet '{sname}': FAILED — {ex}")
                wb.close()

                _add(
                    f"Images Uploaded ({len(uploaded_images)} of {len(image_log)} found)",
                    "text",
                    "\n".join(image_log) if image_log else "(no embedded images found)",
                )

                prompt_text = get_prompt("extract_all_file").replace("{sheet_hint}", sheet_hint)
                _add("System Prompt", "prompt", get_prompt("system"))
                _add("Final Prompt (with hints)", "prompt", prompt_text)

                model = genai.GenerativeModel(
                    model_name=s.gemini_model,
                    system_instruction=get_prompt("system"),
                    generation_config=genai.GenerationConfig(
                        response_mime_type="application/json",
                        temperature=0.1,
                    ),
                )
                response = model.generate_content([*uploaded_images, sheet_text, prompt_text])
                raw = response.text
                _add("Raw AI Response", "json_raw", raw)

                result = _parse_llm_json(raw)
                if isinstance(result, list):
                    result = {"flows": result}
                _add("Parsed Result", "json", result)

            finally:
                os.unlink(tmp_path)

        else:
            # Non-XLSX: convert via markitdown then use text extraction
            from markitdown import MarkItDown
            md_result = MarkItDown().convert(_io.BytesIO(file_bytes))
            markdown = md_result.text_content
            _add("Document Markdown", "markdown", markdown)

            prompt_text = get_prompt("extract_all").replace("{content}", markdown)
            _add("System Prompt", "prompt", get_prompt("system"))
            _add("Final Prompt", "prompt", prompt_text)

            model = genai.GenerativeModel(
                model_name=s.gemini_model,
                system_instruction=get_prompt("system"),
                generation_config=genai.GenerationConfig(
                    response_mime_type="application/json",
                    temperature=0.1,
                ),
            )
            response = model.generate_content(prompt_text)
            raw = response.text
            _add("Raw AI Response", "json_raw", raw)

            result = _parse_llm_json(raw)
            if isinstance(result, list):
                result = {"flows": result}
            _add("Parsed Result", "json", result)

        return {"steps": steps, "error": None}

    except Exception as e:
        _add("Error", "error", traceback.format_exc())
        return {"steps": steps, "error": str(e)}


def extract_document_metadata(full_markdown: str, filename: str) -> dict:
    """Extract high-level document metadata from the full markdown."""
    _init_client()
    s = get_settings()
    prompt = (
        get_prompt("metadata")
        .replace("{filename}", filename)
        .replace("{content}", full_markdown[:3000])
    )

    try:
        model = genai.GenerativeModel(
            model_name=s.gemini_model,
            system_instruction=get_prompt("system"),
            generation_config=genai.GenerationConfig(
                response_mime_type="application/json",
                temperature=0.1,
            ),
        )
        response = model.generate_content(prompt)
        return _parse_llm_json(response.text)
    except Exception:
        return {
            "name": filename,
            "partner_name": None,
            "flow_name": None,
            "version": None,
            "doc_date": None,
        }


# ─── Single-API re-extraction ───────────────────────────────────


def reextract_api(
    api_name: str,
    method: str | None,
    path: str | None,
    exposed_by: str,
    files: list[tuple[bytes, str]],  # (content_bytes, mime_type)
) -> dict:
    """
    Re-extract a single API's spec from uploaded screenshots and/or markdown text.

    files: list of (bytes, mime_type). Images are uploaded to Gemini File API;
           text/markdown is passed inline as text content.

    Returns: { description, method, path, request, response, errors, edge_cases }
    On failure: { "_error": "..." }
    """
    import io

    _init_client()
    s = get_settings()

    prompt = (
        get_prompt("reextract_api")
        .replace("{api_name}", api_name)
        .replace("{method}", method or "unknown")
        .replace("{path}", path or "unknown")
        .replace("{exposed_by}", exposed_by)
    )

    content_parts = []
    text_parts = []

    for file_bytes, mime_type in files:
        if mime_type.startswith("image/"):
            try:
                gfile = genai.upload_file(
                    path=io.BytesIO(file_bytes),
                    mime_type=mime_type,
                    display_name=f"reextract_{api_name}.{mime_type.split('/')[-1]}",
                )
                content_parts.append(gfile)
                logger.info("reextract_api: uploaded image mime=%s", mime_type)
            except Exception:
                logger.warning("reextract_api: failed to upload image", exc_info=True)
        elif mime_type in ("text/plain", "text/markdown"):
            text_parts.append(file_bytes.decode("utf-8", errors="replace"))
        else:
            # Try to decode as text for unknown types
            try:
                text_parts.append(file_bytes.decode("utf-8", errors="replace"))
            except Exception:
                pass

    if text_parts:
        content_parts.append("\n\n---\n\n".join(text_parts))

    content_parts.append(prompt)

    try:
        model = genai.GenerativeModel(
            model_name=s.gemini_model,
            system_instruction=get_prompt("system"),
            generation_config=genai.GenerationConfig(
                response_mime_type="application/json",
                temperature=0.1,
            ),
        )
        logger.info("reextract_api: sending to Gemini api=%s files=%d", api_name, len(files))
        response = model.generate_content(content_parts)
        result = _parse_llm_json(response.text)
        logger.info("reextract_api: success api=%s", api_name)
        return result
    except Exception as e:
        logger.error("reextract_api failed api=%s\n%s", api_name, traceback.format_exc())
        return {"_error": str(e)}
