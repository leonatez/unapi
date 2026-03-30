"""
Step 1: Convert uploaded file → Markdown using markitdown (non-XLSX),
or list sheets for XLSX files (for user-driven sheet selection).

Also provides Gemini File API upload for the new native ingestion path.
"""
import logging
import os
import shutil
from pathlib import Path
from markitdown import MarkItDown
from app.core.config import get_settings
import google.generativeai as genai

logger = logging.getLogger(__name__)

_md = MarkItDown()
_gemini_initialized = False


def _init_gemini():
    global _gemini_initialized
    if not _gemini_initialized:
        s = get_settings()
        genai.configure(api_key=s.gemini_api_key)
        _gemini_initialized = True


def ingest_file(upload_path: str) -> str:
    """
    Convert any supported file to Markdown text (markitdown path).
    Returns the full Markdown string.
    """
    result = _md.convert(upload_path)
    return result.text_content


def save_upload(file_bytes: bytes, filename: str) -> str:
    """Persist the uploaded file and return its local path."""
    settings = get_settings()
    dest = os.path.join(settings.upload_dir, filename)
    with open(dest, "wb") as f:
        f.write(file_bytes)
    return dest


def list_xlsx_sheets(file_path: str) -> list[dict]:
    """
    Read an XLSX file and return metadata for each sheet:
      { name, row_count, col_count, preview: [[cell, ...], ...] }
    Preview contains up to 6 rows (headers + first 5 data rows).
    Uses read_only mode to avoid loading images into memory.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        preview_rows = []
        row_count = 0
        col_count = 0
        for row in ws.iter_rows(values_only=True):
            # Skip fully empty rows
            if not any(cell is not None for cell in row):
                continue
            row_count += 1
            col_count = max(col_count, len(row))
            if len(preview_rows) < 6:
                preview_rows.append([str(c) if c is not None else "" for c in row])
        sheets.append({
            "name": sheet_name,
            "row_count": row_count,
            "col_count": col_count,
            "preview": preview_rows,
        })
    wb.close()
    return sheets


def xlsx_to_markdown(file_path: str, selected_sheets: list[str] | None = None) -> str:
    """
    Convert selected sheets of an XLSX file to markdown suitable for LLM consumption.

    Improvements over the naive flat-table approach:
    - Merged cells: value shown only at the top-left cell; continuation cells are empty
      (avoids duplicating the same label across N columns)
    - Multiple table regions per sheet: detected by finding contiguous non-empty row blocks
      and rendered as separate markdown tables
    - Sparse / single-column rows: rendered as prose or key-value pairs instead of
      broken tables
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    output = []
    sheet_names = selected_sheets if selected_sheets else wb.sheetnames

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        output.append(f"\n## Sheet: {sheet_name}\n")

        # ── Resolve merged cells ──────────────────────────────────────────
        # Only the top-left cell of each merge keeps its value.
        # All continuation cells are treated as empty to prevent duplication.
        merge_top_left_values: dict[tuple, object] = {}
        merge_continuations: set[tuple] = set()
        for merge in ws.merged_cells.ranges:
            top_left_val = ws.cell(merge.min_row, merge.min_col).value
            merge_top_left_values[(merge.min_row, merge.min_col)] = top_left_val
            for r in range(merge.min_row, merge.max_row + 1):
                for c in range(merge.min_col, merge.max_col + 1):
                    if not (r == merge.min_row and c == merge.min_col):
                        merge_continuations.add((r, c))

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        def get_val(r: int, c: int):
            if (r, c) in merge_continuations:
                return None
            if (r, c) in merge_top_left_values:
                return merge_top_left_values[(r, c)]
            return ws.cell(r, c).value

        def cell_str(v) -> str:
            if v is None:
                return ""
            return str(v).strip().replace("\n", " ")

        # ── Build 2-D grid ───────────────────────────────────────────────
        grid: list[list[str]] = []
        for r in range(1, max_row + 1):
            grid.append([cell_str(get_val(r, c)) for c in range(1, max_col + 1)])

        if not grid:
            output.append("(empty)\n")
            continue

        def row_is_empty(row: list[str]) -> bool:
            return all(v == "" for v in row)

        # ── Detect and render contiguous non-empty row blocks ────────────
        i = 0
        while i < len(grid):
            if row_is_empty(grid[i]):
                i += 1
                continue

            block_start = i
            while i < len(grid) and not row_is_empty(grid[i]):
                i += 1
            block = grid[block_start:i]

            # Find columns that have at least one non-empty value in this block
            used_cols = [c for c in range(max_col) if any(row[c] != "" for row in block)]
            if not used_cols:
                continue

            col_count = sum(1 for v in block[0] if v != "")

            if len(block) >= 2 and col_count >= 2:
                # Render as markdown table
                header_cells = [block[0][c] if c < len(block[0]) else "" for c in used_cols]
                output.append("| " + " | ".join(header_cells) + " |")
                output.append("| " + " | ".join(["---"] * len(used_cols)) + " |")
                for row in block[1:]:
                    cells = [row[c] if c < len(row) else "" for c in used_cols]
                    output.append("| " + " | ".join(cells) + " |")
                output.append("")
            else:
                # Sparse block — render as prose / key-value
                for row in block:
                    non_empty = [v for v in row if v != ""]
                    if len(non_empty) == 1:
                        output.append(non_empty[0])
                    elif len(non_empty) == 2:
                        output.append(f"**{non_empty[0]}:** {non_empty[1]}")
                    else:
                        output.append("  ".join(non_empty))
                output.append("")

        output.append("---\n")

    wb.close()
    return "\n".join(output)


def upload_to_gemini(file_path: str) -> str:
    """
    Upload a file to the Gemini File API.
    Returns the file URI (e.g. 'files/abc123') for use in generate_content calls.
    """
    _init_gemini()
    logger.info("Uploading to Gemini File API: %s", file_path)
    uploaded = genai.upload_file(path=file_path)
    logger.info("Gemini upload complete: uri=%s name=%s", uploaded.uri, uploaded.name)
    return uploaded.uri
